# --- Imports ---
import streamlit as st
import pandas as pd
from groq import Groq
import os
from dotenv import load_dotenv
import pdfplumber
from PIL import Image
import pytesseract
import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl.utils import get_column_letter
import io
from manager import NavigationManager, require_login, apply_sidebar_style, set_background_css

# --- Styling & Auth ---
apply_sidebar_style()
set_background_css()
require_login()

# --- Environment Setup ---
load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")
client = Groq(api_key=groq_api_key) if groq_api_key else None

# --- Firebase Init ---
FIREBASE_JSON = "D:/Resume Analyzer/test-23ffe-cf207eed55fe.json"
if not firebase_admin._apps:
    cred = credentials.Certificate(FIREBASE_JSON)
    firebase_admin.initialize_app(cred)
db = firestore.client()

# --- Output Fields ---
EXPECTED_FIELDS = [
    "Full Name", "Gender", "Phone Number", "Email Address", "Location",
    "Total Experience (in years)", "Most Recent Job Title", "Highest Qualification",
    "Key Skills", "ATS Score"
]

# --- Job Criteria Selection ---
job_criteria = {}

# Priority: Form -> Selected from history
if "prefill_criteria" in st.session_state:
    job_criteria = st.session_state["prefill_criteria"]
elif "selected_job_criteria" in st.session_state:
    job_criteria = st.session_state["selected_job_criteria"]

# Validate presence
if not job_criteria:
    st.warning("Job criteria not found. Please fill the form or select from history.")
    st.stop()

# Store unified for downstream use
st.session_state["job_criteria"] = job_criteria

# --- Title ---
st.title("Resume Parser")

# --- Upload Options ---
upload_option = st.selectbox(
    "Choose Input Option",
    options=["Upload New Resumes", "Upload Existing Excel"],
    index=0
)

# === Upload Excel Section ===
if upload_option == "Upload Existing Excel":
    excel_file = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])
    if excel_file:
        try:
            df = pd.read_excel(excel_file)
            missing = [col for col in EXPECTED_FIELDS if col not in df.columns]
            if missing:
                st.error(f"Missing required columns: {', '.join(missing)}")
            else:
                st.session_state.df = df
                st.session_state.processed = True
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Applicants")
                    worksheet = writer.book["Applicants"]
                    for column_cells in worksheet.columns:
                        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2
                excel_buffer.seek(0)
                st.session_state.excel_data = excel_buffer.read()
                st.success("Excel uploaded successfully.")
                if st.button("Analyze Uploaded Excel"):
                    st.switch_page("pages/Dashboard.py")
        except Exception as e:
            st.error(f"Failed to read Excel file: {e}")

# === Upload Resume Section ===
if upload_option == "Upload New Resumes":
    st.markdown("### Upload Resume Files")
    uploaded_files = st.file_uploader(
        "Upload Resume Files (PDFs or Images)",
        type=["pdf", "png", "jpg", "jpeg"],
        accept_multiple_files=True
    )

    if "processed" not in st.session_state:
        st.session_state["processed"] = False

    if uploaded_files and not st.session_state["processed"]:
        st.info("Processing resumes...")
        rows, failed_files = [], []
        progress_bar = st.progress(0)
        progress_text = st.empty()

        for i, file in enumerate(uploaded_files):
            ext = file.name.split('.')[-1].lower()
            text = extract_text_from_pdf(file) if ext == "pdf" else extract_text_from_image(file)
            if text:
                extracted_info = extract_info_with_groq(text, job_criteria)
                if extracted_info:
                    extracted_info["Resume File"] = file.name
                    rows.append(extracted_info)
                else:
                    failed_files.append(file.name)
            else:
                failed_files.append(file.name)

            progress_bar.progress((i + 1) / len(uploaded_files))
            progress_text.text(f"Processed {i + 1}/{len(uploaded_files)}")

        if rows:
            df = pd.DataFrame(rows, columns=EXPECTED_FIELDS)
            st.session_state.df = df
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Applicants")
                worksheet = writer.book["Applicants"]
                for column_cells in worksheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2
            excel_buffer.seek(0)
            st.session_state.excel_data = excel_buffer.read()
            st.session_state["processed"] = True
            st.success("Parsing complete.")
            if failed_files:
                st.warning(f"Some files failed: {', '.join(failed_files)}")
        else:
            st.warning("No data could be extracted.")

# --- Show Results ---
if st.session_state.get("processed") and "df" in st.session_state:
    st.subheader("Parsed Resume Data")
    st.dataframe(st.session_state.df)

    st.download_button(
        label="Download Excel",
        data=st.session_state.excel_data,
        file_name="aggregated_resumes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if st.button("Analyze Parsed Resumes"):
        st.switch_page("pages/Dashboard.py")

    if st.button("Reset"):
        st.session_state["processed"] = False
        st.session_state.pop("df", None)
        st.session_state.pop("excel_data", None)

# --- Helper: PDF Text ---
def extract_text_from_pdf(file):
    try:
        with pdfplumber.open(file) as pdf:
            return ''.join(page.extract_text() or '' for page in pdf.pages).strip()
    except Exception as e:
        st.error(f"PDF Error ({file.name}): {e}")
        return ""

# --- Helper: Image Text ---
def extract_text_from_image(file):
    try:
        image = Image.open(file).convert("L")
        return pytesseract.image_to_string(image).strip()
    except Exception as e:
        st.error(f"Image Error ({file.name}): {e}")
        return ""

# --- Groq Resume Parser ---
def extract_info_with_groq(resume_text, job_criteria, timeout=90):
    if not client:
        return None

    prompt = f"""
You are a resume evaluation expert. Based on the job criteria below, evaluate the resume and extract each field in the format `Field: Value`.
Use 'N/A' if data is missing. Calculate ATS Score out of 100 based on relevance to role, experience, skills, and location.
Return only these fields (no extra text):
- Full Name
- Gender
- Phone Number
- Email Address
- Location
- Total Experience (in years)
- Most Recent Job Title
- Highest Qualification
- Key Skills
- ATS Score

Job Criteria:
- Role: {job_criteria.get('role', 'N/A')}
- Must-Have Skills: {', '.join(job_criteria.get('must_have_skills', []))}
- Minimum Experience: {job_criteria.get('min_experience', '0')} years
- Preferred Locations: {', '.join(job_criteria.get('preferred_locations', []))}
- Job Description: {job_criteria.get('job_description', '')}

Resume Text:
```{resume_text}```
"""

    try:
        response = client.chat.completions.create(
            model="llama3-8b-8192",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=500,
            timeout=timeout
        )
        output = response.choices[0].message.content.strip()
        extracted = {}
        for line in output.split('\n'):
            if ':' in line:
                key, value = map(str.strip, line.split(':', 1))
                if key in EXPECTED_FIELDS:
                    extracted[key] = value or "N/A"
        for field in EXPECTED_FIELDS:
            extracted.setdefault(field, "N/A")
        return extracted
    except Exception as e:
        st.error(f"Groq error: {e}")
        return None
