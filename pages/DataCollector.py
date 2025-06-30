import streamlit as st
import pandas as pd
from groq import Groq
import os
from dotenv import load_dotenv
import pdfplumber
from PIL import Image
import pytesseract
import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl.utils import get_column_letter
import io
from manager import require_login, apply_sidebar_style, set_background_css, hide_sidebar_pages

# --- Styling & Auth ---
apply_sidebar_style()
hide_sidebar_pages()
set_background_css()
require_login()


# --- Load Environment Variables ---
load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")
client = Groq(api_key=groq_api_key)

# --- Firebase Init ---
FIREBASE_JSON = "D:/Resume Analyzer/test-23ffe-cf207eed55fe.json"
if not firebase_admin._apps:
    cred = credentials.Certificate(FIREBASE_JSON)
    firebase_admin.initialize_app(cred)
db = firestore.client()

# --- Fields to Extract ---
EXPECTED_FIELDS = [
    "Full Name", "Gender", "Phone Number", "Email Address", "Location",
    "Total Experience (in years)", "Most Recent Job Title", "Highest Qualification",
    "Key Skills", "ATS Score", "Remark"
]

# --- Helper Functions ---
def extract_text_from_pdf(file):
    try:
        with pdfplumber.open(file) as pdf:
            return ''.join(page.extract_text() or '' for page in pdf.pages).strip()
    except Exception as e:
        st.error(f"PDF Error ({file.name}): {e}")
        return ""

def extract_text_from_image(file):
    try:
        image = Image.open(file).convert("L")
        return pytesseract.image_to_string(image).strip()
    except Exception as e:
        st.error(f"Image Error ({file.name}): {e}")
        return ""

def extract_info_with_groq(resume_text, timeout=90):
    if not client:
        return None

    prompt = f"""
Extract the following information from the resume text provided below. 
Return each field exactly in this format: Field: Value. 
the resume text is extracted from a pdf and may be in random order, every information is provided.
the following data is to be passed to a parser to be automatically input into an excel, make sure that integrity is maintained by no excess text in the output.
the response should be'
Full Name:
Gender:
Phone Number:
Email Address:
Location:
Total Experience:
Most Recent Job Title:
Highest Qualification:
Skills:
ATS Score:
'only these fields should be filled in response,
Assume gender based on name if not mentioned,
calculate total experience other wise put 0 by default only give number of years no other text,
calculate ats score based on 
    Job Criteria:
    Role: {job_criteria.get('role', '')}
    Job Description: {job_criteria.get('job_description', '')}
    Must-Have Skills: {', '.join(job_criteria.get('must_have_skills', []))}
    Min Experience: {job_criteria.get('min_experience', '')}
    ATS score is based on how well the resume match the criteria.
    if any of the data is assumed dont mention it nor write anything in '()'
example:
'
Full Name: john doe
Gender: Male
Phone Number: 9856428597
Email Address: johndoe@mail.com
Location: New York
Total Experience: 8
Most Recent Job Title: Data Scientist
Highest Qualification: Batchlor in Computer Science
Skills: Apache, DBMS, Collab, Python
ATS Score: 90
Remark:
'
In the Remark section add explaination whioch was omitted in other fields, 
make sure to not give remark about the field in the field inself.
Resume Text:
```{resume_text}```
"""
    try:
        response = client.chat.completions.create(
            model="llama3-8b-8192",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=500,
            timeout=timeout
        )
        output = response.choices[0].message.content.strip()

        extracted = {}
        for line in output.split('\n'):
            if ':' in line:
                key, value = map(str.strip, line.split(':', 1))
                for expected in EXPECTED_FIELDS:
                    if key.lower() in expected.lower():
                        extracted[expected] = value or "N/A"
                        break
        for field in EXPECTED_FIELDS:
            extracted.setdefault(field, "N/A")
        return extracted

    except Exception as e:
        st.error(f"Groq error: {e}")
        return None

# --- Page Title ---
st.title("Resume Parser")

# --- Get Job Criteria from Session ---
job_criteria = (
    st.session_state.get("current_criteria") or
    st.session_state.get("prefill_criteria") or
    st.session_state.get("selected_job_criteria")
)


if not job_criteria:
    st.warning("⚠️ Job criteria not found. Please fill the form or select from your history.")
    st.stop()
    
    
# --- Optional Excel Upload ---
with st.expander("📂 Already have parsed data? Upload Excel"):
    excel_file = st.file_uploader("Upload Excel", type=["xlsx"])
    
    if excel_file:
        try:
            df = pd.read_excel(excel_file)
            missing = [col for col in EXPECTED_FIELDS if col not in df.columns]
            
            if missing:
                st.error(f"Missing columns: {', '.join(missing)}")
            else:
                # Clear parsed resume data to keep only one source
                for k in ["analysis_source", "df", "excel_data"]:
                    st.session_state.pop(k, None)

                # Save new dataframe to session
                st.session_state.df = df
                st.session_state.processed = True

                # Save formatted Excel to buffer
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Applicants")
                    sheet = writer.book["Applicants"]
                    for col in sheet.columns:
                        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
                buffer.seek(0)
                st.session_state.excel_data = buffer.read()

                st.success("✅ Excel uploaded and validated.")
        
        except Exception as e:
            st.error(f"❌ Excel error: {e}")

    # Show Analyze button only if Excel is uploaded and valid
    if "df" in st.session_state and st.session_state.get("processed"):
        if st.button("📊 Analyze Uploaded Excel"):
            st.session_state.analysis_source = "uploaded_excel"
            st.switch_page("pages/Dashboard.py")

# --- Upload Resumes ---
st.markdown("### 📄 Upload Resume Files")
files = st.file_uploader("Upload PDFs or Images", type=["pdf", "png", "jpg", "jpeg"], accept_multiple_files=True)

if "processed" not in st.session_state:
    st.session_state["processed"] = False

if files and not st.session_state["processed"]:
    st.info("Processing resumes...")

    rows, failed = [], []
    progress_bar = st.progress(0)
    status = st.empty()

    for i, f in enumerate(files):
        ext = f.name.lower().split('.')[-1]
        text = extract_text_from_pdf(f) if ext == "pdf" else extract_text_from_image(f)

        if text:
            info = extract_info_with_groq(text)
            if info:
                info["Resume File"] = f.name
                rows.append(info)
            else:
                failed.append(f.name)
        else:
            failed.append(f.name)

        progress_bar.progress((i + 1) / len(files))
        status.text(f"Processed {i + 1}/{len(files)}")

    if rows:
        df = pd.DataFrame(rows, columns=EXPECTED_FIELDS)
        st.session_state.df = df

        # Save to Excel
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Applicants")
            worksheet = writer.book["Applicants"]
            for column_cells in worksheet.columns:
                max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2
        excel_buffer.seek(0)
        st.session_state.excel_data = excel_buffer.read()
        st.session_state["processed"] = True

        st.success(f"✅ Parsed {len(rows)} resumes successfully.")
        if failed:
            st.warning(f"⚠️ Failed to process: {', '.join(failed)}")
    else:
        st.warning("No valid data extracted.")

# --- Display Results ---
if st.session_state.get("processed") and "df" in st.session_state:
    st.subheader("📊 Parsed Resume Data")
    st.dataframe(st.session_state.df)

    st.download_button("📥 Download Excel", st.session_state.excel_data, "aggregated_resumes.xlsx")

    # Analyze Parsed Resumes
    if st.button("📊 Analyze Parsed Resumes"):
        # Ensure only parsed resume data is carried forward
        for key in ["analysis_source"]:
            st.session_state.pop(key, None)

        st.session_state.analysis_source = "parsed_resumes"
        st.switch_page("pages/Dashboard.py")

    # Reset all session state related to resume parsing
    if st.button("🔄 Reset"):
        for k in ["processed", "df", "excel_data", "analysis_source"]:
            st.session_state.pop(k, None)

